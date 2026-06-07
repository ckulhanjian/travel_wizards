#!/usr/bin/env python3
"""
Hotel Invoice Processor — Travel Wizards
"""

import os
import sys
import re
import shutil
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

# Editor is imported lazily to avoid circular issues and keep startup fast

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError as e:
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Missing Dependencies",
                         f"Missing required library: {e}\n\n"
                         "Please install:\n"
                         "pip install pandas openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Asset path helper (works as .py or PyInstaller bundle)
# ---------------------------------------------------------------------------
def _asset(filename):
    if getattr(sys, "frozen", False):
        base = sys._MEIPASS
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, filename)


# ---------------------------------------------------------------------------
# Template registry  — add more entries here when additional templates exist
# ---------------------------------------------------------------------------
TEMPLATES = {
    "Hotel Invoice  (standard)": _asset("Hotel_Invoice-Template.xlsx"),
    # "Hotel Invoice  (international)": _asset("Hotel_Invoice-Template-Intl.xlsx"),
    # "Hotel Invoice  (group)":         _asset("Hotel_Invoice-Template-Group.xlsx"),
}


# ---------------------------------------------------------------------------
# Parse Hotels report (all sheets)
# ---------------------------------------------------------------------------
def load_hotels_data(hotels_path: str) -> "pd.DataFrame":
    df = pd.read_excel(hotels_path, sheet_name=0, header=None)
    rows = []

    # Find the header row (contains "Date")
    header_row_idx = None
    for i, row in df.iterrows():
        if "Date" in [str(v).strip() for v in row.values]:
            header_row_idx = i
            break
    if header_row_idx is None:
        return pd.DataFrame(rows)

    header = list(df.iloc[header_row_idx])
    col_map = {}
    for ci, h in enumerate(header):
        hs = str(h).strip()
        if hs == "Date":        col_map["Date"]      = ci
        elif hs == "Account":   col_map["Account"]   = ci
        elif hs == "Invoice":   col_map["Invoice"]   = ci
        elif hs == "Traveler":  col_map["Traveler"]  = ci
        elif hs == "Itinerary": col_map["Itinerary"] = ci
        elif hs == "Total":     col_map["Total"]     = ci
        elif hs == "Comm":      col_map["Comm"]      = ci
        elif hs == "Depart":    col_map["Depart"]    = ci

    if not col_map:
        return pd.DataFrame(rows)

    # Only process the 39 data rows immediately after the header
    # (Excel rows 9–47). This excludes Grand Total and summary rows.
    data_end = header_row_idx + 40  # header+1 through header+39 inclusive
    for i in range(header_row_idx + 1, min(data_end, len(df))):
        row = df.iloc[i]
        if "R T H P" not in str(row.iloc[0]).strip():
            continue

        def get(key, offset=0):
            if key not in col_map:
                return ""
            ci = col_map[key] + offset
            if ci < len(row):
                v = row.iloc[ci]
                return "" if pd.isna(v) else str(v).strip()
            return ""

        # Hotel name can overflow into adjacent columns
        hotel_parts = [get("Itinerary")]
        for extra in range(1, 4):
            part = get("Itinerary", extra)
            if (part and part not in ("NaN", "nan")
                    and not re.match(r"^\d+\.\d+$", part)
                    and not re.match(r"^\d{2}/\d{2}", part)):
                hotel_parts.append(part)
            else:
                break
        hotel_name = " ".join(p for p in hotel_parts if p and p != "NaN")

        date_raw = get("Date")
        date_m = re.search(r"\d{2}/\d{2}/\d{2,4}", date_raw)
        date_str = date_m.group(0) if date_m else date_raw

        account = get("Account")
        if not account or account in ("NaN", "nan"):
            agent_m = re.search(r"[A-Z]{2,}", date_raw)
            account = agent_m.group(0) if agent_m else ""

        rows.append({
            "Date":     date_str,
            "Account":  account.strip(),
            "Invoice":  get("Invoice"),
            "Traveler": get("Traveler").strip(),
            "Hotel":    hotel_name,
            "Total":    get("Total"),
            "Comm":     get("Comm"),
            "Depart":   get("Depart"),
        })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Name helpers
# ---------------------------------------------------------------------------
def format_guest_name(traveler: str) -> str:
    t = traveler.strip()
    if "/" in t:
        last, first = t.split("/", 1)
        return f"{first.strip()} {last.strip()}"
    return t


def last_name_only(traveler: str) -> str:
    if "/" in traveler:
        return traveler.split("/")[0].strip()
    return traveler.strip()


# ---------------------------------------------------------------------------
# Fill one invoice from template
# ---------------------------------------------------------------------------
BLACK      = "FF000000"
TNR        = "Times New Roman"
MONEY_FMT  = '"$"#,##0.00'


def _set(ws, coord, value, bold=None, number_format=None, align_right=False):
    """Write a value with Times New Roman 11pt black, preserving bold from template."""
    from openpyxl.styles import Font, Alignment
    cell = ws[coord]
    cell.value = value
    existing = cell.font
    cell.font = Font(
        name=TNR,
        size=11,
        bold=existing.bold if bold is None else bold,
        color=BLACK,
    )
    if number_format:
        cell.number_format = number_format
    if align_right:
        cell.alignment = Alignment(horizontal="right")


def fill_invoice(template_path: str, row: dict, output_path: str):
    """Fill one invoice from a single data row."""
    from openpyxl.styles import Font
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # Force ALL existing template cells to Times New Roman 11pt black
    for r in ws.iter_rows():
        for cell in r:
            if cell.value is not None:
                existing = cell.font
                cell.font = Font(
                    name=TNR,
                    size=11,
                    bold=existing.bold,
                    color=BLACK,
                )

    # Header fields
    _set(ws, "D3", row["Invoice"].lstrip("0"), align_right=True)
    _set(ws, "D4", row["Date"], align_right=True)
    _set(ws, "D5", row["Account"], align_right=True)

    # Hotel name on the row BELOW the "HOTEL:" label (A17)
    _set(ws, "A17", row["Hotel"].upper() if row["Hotel"] else "")

    # Guest name on the row BELOW "Guest(s):" label (A24)
    _set(ws, "A24", format_guest_name(row["Traveler"]))

    # Depart date
    _set(ws, "D26", row["Depart"], align_right=True)

    # Financials — stored as numbers, formatted as $#,##0.00
    try:
        total = float(row["Total"]) if row["Total"] else 0.0
        comm  = float(row["Comm"])  if row["Comm"]  else 0.0
    except (ValueError, TypeError):
        total = comm = 0.0

    subtotal = total - comm

    _set(ws, "D28", total,    number_format=MONEY_FMT, align_right=True)
    _set(ws, "D29", comm,     number_format=MONEY_FMT, align_right=True)
    _set(ws, "D31", subtotal, number_format=MONEY_FMT, align_right=True)
    _set(ws, "D35", subtotal, number_format=MONEY_FMT, align_right=True, bold=True)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _center_window(win, w, h):
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------
class HotelInvoiceGUI:
    CLR_BG      = "#ffffff"
    CLR_PANEL   = "#f5f5f5"
    CLR_BORDER  = "#cccccc"
    CLR_MUTED   = "#555555"
    CLR_TEXT    = "#000000"
    CLR_LOG_BG  = "#f9f9f9"
    CLR_LOG_FG  = "#111111"
    CLR_BTN     = "#e0e0e0"
    CLR_BTN_ACT = "#cccccc"

    def __init__(self, parent=None):
        if parent is None:
            self.root = tk.Tk()
        else:
            self.root = tk.Toplevel(parent)
        self.root.title("Travel Wizards — Hotel Invoice Processor")
        self.root.resizable(True, True)
        self.root.configure(bg=self.CLR_BG)
        _center_window(self.root, 900, 580)
        self.root.option_add("*Button.Background",       self.CLR_BTN)
        self.root.option_add("*Button.Foreground",       "#000000")
        self.root.option_add("*Button.activeBackground", self.CLR_BTN_ACT)
        self.root.option_add("*Button.activeForeground", "#000000")
        self.root.option_add("*Button.relief",           "flat")

        self.hotels_file    = tk.StringVar()
        self.template_label = tk.StringVar(value=list(TEMPLATES.keys())[0])
        self.stats_var      = tk.StringVar(value="")

        self._draw_logo()
        self._setup_ui()

    # ------------------------------------------------------------------
    def _draw_logo(self):
        c = tk.Canvas(self.root, height=74, bg=self.CLR_BG,
                      highlightthickness=0, bd=0)
        c.pack(fill="x")
        c.create_line(0, 73, 2000, 73, fill=self.CLR_BORDER, width=1)
        c.create_text(450, 26, text="TRAVEL  WIZARDS",
                      font=("Georgia", 26, "bold"),
                      fill="#000000", anchor="center")
        c.create_text(450, 52,
                      text="H O T E L   I N V O I C E   P R O C E S S O R",
                      font=("Arial", 10), fill="#cc0000", anchor="center")
        if isinstance(self.root, tk.Toplevel):
            tk.Button(self.root, text="⌂  Home", command=self.root.destroy,
                      relief="flat", cursor="hand2",
                      bg=self.CLR_BG, fg="#000000",
                      activebackground=self.CLR_BG,
                      font=("Arial", 9, "bold"),
                      padx=10, pady=4, bd=0).place(x=10, y=10)

    # ------------------------------------------------------------------
    def _path_row(self, parent, label, var, is_folder=False):
        outer = tk.Frame(parent, bg=self.CLR_BG)
        outer.pack(fill="x", padx=24, pady=(10, 0))
        tk.Label(outer, text=label, font=("Arial", 10, "bold"),
                 bg=self.CLR_BG, fg=self.CLR_MUTED).pack(anchor="w")
        box = tk.Frame(outer, bg=self.CLR_PANEL,
                       highlightbackground=self.CLR_BORDER,
                       highlightthickness=1)
        box.pack(fill="x", pady=(4, 0))
        tk.Entry(box, textvariable=var, relief="flat",
                 bg=self.CLR_PANEL, fg=self.CLR_TEXT,
                 readonlybackground=self.CLR_PANEL,
                 insertbackground=self.CLR_TEXT,
                 font=("Consolas", 11), bd=0,
                 state="readonly").pack(side="left", padx=10, pady=8,
                                        fill="x", expand=True)
        cmd = (lambda v=var: self._browse_folder(v)) if is_folder \
              else (lambda v=var: self._browse_file(v))
        tk.Button(box, text="Browse", command=cmd,
                  relief="flat", cursor="hand2",
                  bg=self.CLR_BTN, fg="#000000",
                  activebackground=self.CLR_BTN_ACT,
                  font=("Arial", 10, "bold"),
                  padx=14, pady=6, bd=0).pack(side="right", padx=6, pady=4)

    def _template_row(self, parent):
        outer = tk.Frame(parent, bg=self.CLR_BG)
        outer.pack(fill="x", padx=24, pady=(10, 0))
        tk.Label(outer, text="INVOICE TEMPLATE", font=("Arial", 10, "bold"),
                 bg=self.CLR_BG, fg=self.CLR_MUTED).pack(anchor="w")
        box = tk.Frame(outer, bg=self.CLR_PANEL,
                       highlightbackground=self.CLR_BORDER,
                       highlightthickness=1)
        box.pack(fill="x", pady=(4, 0))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TW.TCombobox",
                        fieldbackground=self.CLR_PANEL,
                        background=self.CLR_PANEL,
                        foreground=self.CLR_TEXT,
                        selectbackground=self.CLR_PANEL,
                        selectforeground=self.CLR_TEXT,
                        borderwidth=0,
                        relief="flat")
        style.map("TW.TCombobox",
                  fieldbackground=[("readonly", self.CLR_PANEL)],
                  selectbackground=[("readonly", self.CLR_PANEL)],
                  selectforeground=[("readonly", self.CLR_TEXT)])

        cb = ttk.Combobox(box,
                          textvariable=self.template_label,
                          values=list(TEMPLATES.keys()),
                          state="readonly",
                          style="TW.TCombobox",
                          font=("Consolas", 11))
        cb.pack(side="left", padx=10, pady=8, fill="x", expand=True)

    def _browse_file(self, var):
        path = filedialog.askopenfilename(
            title="Select file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if path:
            var.set(path)

    def _browse_folder(self, var):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            var.set(path)

    # ------------------------------------------------------------------
    def _setup_ui(self):
        self._path_row(self.root, "HOTELS REPORT  (.xlsx)", self.hotels_file)
        self._template_row(self.root)

        sf = tk.Frame(self.root, bg=self.CLR_BG)
        sf.pack(fill="x", padx=24, pady=(8, 0))
        tk.Label(sf, textvariable=self.stats_var, font=("Arial", 9),
                 bg=self.CLR_BG, fg=self.CLR_MUTED).pack(anchor="w")

        bf = tk.Frame(self.root, bg=self.CLR_BG)
        bf.pack(fill="x", padx=24, pady=(12, 8))
        self.process_btn = tk.Button(
            bf, text="▶  GENERATE INVOICES",
            command=self.start_processing,
            relief="flat", cursor="hand2",
            bg=self.CLR_BTN, fg="#000000",
            activebackground=self.CLR_BTN_ACT,
            font=("Arial", 13, "bold"),
            pady=10, bd=0)
        self.process_btn.pack(side="left", fill="x", expand=True, padx=(0, 8))

        tk.Button(bf, text="✏  Open Editor",
                  command=self._open_editor,
                  relief="flat", cursor="hand2",
                  bg=self.CLR_BTN, fg="#000000",
                  activebackground=self.CLR_BTN_ACT,
                  font=("Arial", 13, "bold"),
                  pady=10, padx=20, bd=0).pack(side="left")

        tk.Label(self.root, text="PROGRESS LOG", font=("Arial", 10, "bold"),
                 bg=self.CLR_BG, fg=self.CLR_MUTED
                 ).pack(anchor="w", padx=24, pady=(6, 2))
        lo = tk.Frame(self.root, highlightbackground=self.CLR_BORDER,
                      highlightthickness=1, bg=self.CLR_BORDER)
        lo.pack(fill="both", expand=True, padx=24, pady=(0, 16))
        self.log_text = scrolledtext.ScrolledText(
            lo, relief="flat", bd=0,
            bg=self.CLR_LOG_BG, fg=self.CLR_LOG_FG,
            insertbackground=self.CLR_LOG_FG,
            font=("Consolas", 11), wrap="word")
        self.log_text.pack(fill="both", expand=True, padx=1, pady=1)

    # ------------------------------------------------------------------
    def log(self, msg):
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def start_processing(self):
        if not self.hotels_file.get():
            messagebox.showerror("Error", "Please select the Hotels report file.")
            return

        template_path = TEMPLATES.get(self.template_label.get())
        if not template_path or not os.path.exists(template_path):
            messagebox.showerror(
                "Template Not Found",
                f"Template file not found:\n{template_path}\n\n"
                "Place the template .xlsx in the same folder as this script.")
            return

        self.process_btn.config(state="disabled",
                                 text="⏳  Processing...", bg="#aaaaaa")
        self.log_text.delete(1.0, tk.END)
        t = threading.Thread(target=self._process)
        t.daemon = True
        t.start()

    # ------------------------------------------------------------------
    def _process(self):
        try:
            hotels_path   = self.hotels_file.get()
            template_path = TEMPLATES[self.template_label.get()]

            # Save processed invoices alongside the Hotels report file
            output_dir = os.path.join(os.path.dirname(hotels_path), "processed_invoices")
            os.makedirs(output_dir, exist_ok=True)
            self.log(f"Template  : {os.path.basename(template_path)}")
            self.log(f"Output    : {output_dir}\n")

            self.log("Loading Hotels report...")
            df = load_hotels_data(hotels_path)
            self.log(f"Found {len(df)} data rows.")

            if df.empty:
                self.log("⚠  No data rows found. Verify the Hotels file format.")
                return

            total_rows = len(df)
            self.log(f"Rows to process: {total_rows}\n")
            self.stats_var.set(f"Rows found: {total_rows}")

            successful = failed = skipped = 0
            used_names = {}  # track filename collisions → append suffix

            for _, row in df.iterrows():
                invoice_no = str(row.get("Invoice", "")).strip()
                if not invoice_no or invoice_no in ("nan", "NaN", ""):
                    skipped += 1
                    continue

                last = last_name_only(row.get("Traveler", ""))
                safe = re.sub(r'[\\/*?:"<>|]', "_", invoice_no.lstrip("0"))
                base = f"{safe} {last}" if last else safe

                # Avoid overwriting if same invoice# appears on multiple rows
                count = used_names.get(base, 0)
                used_names[base] = count + 1
                fname = f"{base}.xlsx" if count == 0 else f"{base} ({count}).xlsx"
                out_path = os.path.join(output_dir, fname)

                self.log(f"[{invoice_no.lstrip('0')}]  {last}  →  {fname}")
                try:
                    fill_invoice(template_path, row.to_dict(), out_path)
                    self.log("  ✓  Saved")
                    successful += 1
                except Exception as e:
                    self.log(f"  ✗  Error: {e}")
                    failed += 1

            self.log(f"\n{'='*55}")
            self.log("SUMMARY:")
            self.log(f"  Successfully generated : {successful}")
            self.log(f"  Failed                 : {failed}")
            self.log(f"  Skipped (no invoice #) : {skipped}")
            self.log(f"  Output folder          : {output_dir}")
            self.stats_var.set(
                f"Done — {successful} invoices generated"
                + (f", {failed} failed" if failed else ""))

            if successful > 0:
                self.root.after(0, lambda d=output_dir, s=successful, f=failed:
                    self._show_complete_dialog(d, s, f))

        except Exception as e:
            import traceback
            self.log(f"\nFatal error: {e}")
            self.log(traceback.format_exc())
            messagebox.showerror("Error", f"An error occurred:\n{e}")
        finally:
            self.process_btn.config(state="normal",
                                     text="▶  GENERATE INVOICES",
                                     bg=self.CLR_BTN)

    def _show_complete_dialog(self, output_dir, successful, failed):
        """Show completion dialog with option to open the editor."""
        win = tk.Toplevel(self.root)
        win.title("Processing Complete")
        win.geometry("420x220")
        win.resizable(False, False)
        win.configure(bg="#ffffff")
        win.grab_set()

        tk.Label(win, text="✓  Processing Complete",
                 font=("Georgia", 15, "bold"),
                 bg="#ffffff", fg="#000000").pack(pady=(22, 4))
        tk.Label(win,
                 text=f"Generated: {successful} invoice(s)   Failed: {failed}\n"
                      f"Saved in: {os.path.basename(output_dir)}",
                 font=("Arial", 10), bg="#ffffff", fg="#555555",
                 justify="center").pack(pady=(0, 18))

        btn_row = tk.Frame(win, bg="#ffffff")
        btn_row.pack()

        tk.Button(btn_row, text="Open Editor",
                  command=lambda: [win.destroy(), self._open_editor(output_dir)],
                  relief="flat", cursor="hand2",
                  bg="#e0e0e0", fg="#000000",
                  activebackground="#cccccc", activeforeground="#000000",
                  font=("Arial", 11, "bold"),
                  padx=18, pady=8, bd=0).pack(side="left", padx=(0, 12))

        tk.Button(btn_row, text="Close",
                  command=win.destroy,
                  relief="flat", cursor="hand2",
                  bg="#e0e0e0", fg="#000000",
                  activebackground="#cccccc",
                  font=("Arial", 11),
                  padx=18, pady=8, bd=0).pack(side="left")

    def _open_editor(self, output_dir=None):
        """Launch the invoice editor in queue mode over all processed invoices."""
        try:
            from hotel_invoice_editor import InvoiceEditorWindow
        except ImportError:
            messagebox.showerror("Editor Not Found",
                                 "hotel_invoice_editor.py must be in the same "
                                 "folder as this script.")
            return
        queue = []
        if output_dir and os.path.isdir(output_dir):
            queue = sorted(
                os.path.join(output_dir, f)
                for f in os.listdir(output_dir)
                if f.lower().endswith(".xlsx")
            )
        InvoiceEditorWindow(parent=self.root, queue=queue)

    def run(self):
        if isinstance(self.root, tk.Tk):
            self.root.mainloop()


if __name__ == "__main__":
    app = HotelInvoiceGUI()
    app.run()
