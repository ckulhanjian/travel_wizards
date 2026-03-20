#!/usr/bin/env python3
"""
Hotel Invoice Editor — Travel Wizards

Two modes:
  Queue mode  — pass a list of .xlsx files (launched from processor).
                Shows file N of N, prev/next arrows, save-and-next.
  Single mode — browse to one file (standalone).
"""

import os
import re
import sys
import shutil
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    import fitz
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from invoice_pdf import build_pdf
except ImportError as e:
    _r = tk.Tk(); _r.withdraw()
    messagebox.showerror("Missing Dependencies",
                         f"Missing library: {e}\n\npip install PyMuPDF openpyxl reportlab")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Asset / config helpers
# ---------------------------------------------------------------------------
def _asset(filename):
    base = sys._MEIPASS if getattr(sys, "frozen", False) \
           else os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, filename)





BLACK     = "FF000000"
TNR       = "Times New Roman"
MONEY_FMT = '"$"#,##0.00'

FIELDS = [
    # label,               cell,  editable, money,  multiline
    ("Invoice #",          "D3",  True,     False,  False),
    ("Date",               "D4",  True,     False,  False),
    ("Account",            "D5",  True,     False,  False),
    ("PNR Locator",        "D7",  True,     False,  False),
    ("Hotel Name",         "A17", True,     False,  False),
    ("Hotel Address",      "A18", True,     False,  False),
    ("Hotel City/State",   "A19", True,     False,  False),
    ("Hotel Phone",        "A20", True,     False,  False),
    ("Confirmation #",     "D23", True,     False,  False),
    ("Guest(s)",           "A24", True,     False,  True),
    ("Arrive",             "D25", True,     False,  False),
    ("Depart",             "D26", True,     False,  False),
    ("Rate",               "D27", True,     True,   False),
    ("Total",              "D28", True,     True,   False),
    ("Commission",         "D29", True,     True,   False),
    ("Pd. To Date",        "D30", True,     True,   False),
]


# ---------------------------------------------------------------------------
# Read / write xlsx
# ---------------------------------------------------------------------------
def read_fields(xlsx_path: str) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    return {cell: ("" if ws[cell].value is None else str(ws[cell].value))
            for _, cell, *__ in FIELDS}


def write_fields(src: str, values: dict, dst: str):
    """Write updated values to dst. Handles src == dst safely via temp file."""
    import tempfile
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(os.path.abspath(dst)))
    os.close(tmp_fd)
    try:
        shutil.copy2(src, tmp_path)
        wb = load_workbook(tmp_path)
        ws = wb.active

        for label, cell, editable, is_money, multiline in FIELDS:
            raw = values.get(cell, "")
            if is_money:
                try:
                    v = float(raw.replace("$", "").replace(",", "")) if raw else 0.0
                except ValueError:
                    v = 0.0
            else:
                v = raw
            c = ws[cell]
            c.value = v
            c.font = Font(name=TNR, size=11, bold=c.font.bold, color=BLACK)
            if is_money:
                c.number_format = MONEY_FMT
                c.alignment = Alignment(horizontal="right")

        try:
            total    = float(values.get("D28", "0").replace("$", "").replace(",", "") or 0)
            comm     = float(values.get("D29", "0").replace("$", "").replace(",", "") or 0)
            pd_date  = float(values.get("D30", "0").replace("$", "").replace(",", "") or 0)
            subtotal = total - comm - pd_date
        except ValueError:
            subtotal = 0.0

        for coord in ("D31", "D35"):
            c = ws[coord]
            c.value = subtotal
            c.number_format = MONEY_FMT
            c.alignment = Alignment(horizontal="right")
            c.font = Font(name=TNR, size=11, bold=(coord == "D35"), color=BLACK)

        wb.save(tmp_path)
        shutil.move(tmp_path, dst)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise



# ---------------------------------------------------------------------------
# Editor window
# ---------------------------------------------------------------------------
class InvoiceEditorWindow:
    CLR_BG      = "#ffffff"
    CLR_PANEL   = "#f5f5f5"
    CLR_BORDER  = "#cccccc"
    CLR_MUTED   = "#555555"
    CLR_TEXT    = "#000000"
    CLR_BTN     = "#e0e0e0"
    CLR_BTN_ACT = "#cccccc"
    CLR_SECTION = "#efefef"

    def __init__(self, parent=None, queue=None, initial_path=None):
        """
        queue        : list of .xlsx paths for queue mode (from processor)
        initial_path : single file for standalone browse mode
        """
        if parent is None:
            self.root = tk.Tk()
            self.standalone = True
        else:
            self.root = tk.Toplevel(parent)
            self.standalone = False

        self.root.title("Travel Wizards — Invoice Editor")
        self.root.geometry("800x840")
        self.root.resizable(True, True)
        self.root.configure(bg=self.CLR_BG)

        # Queue state
        self._queue   = list(queue) if queue else []
        self._q_index = 0

        self.xlsx_path    = tk.StringVar()
        self.status_var   = tk.StringVar(value="")
        self.counter_var  = tk.StringVar(value="")
        self.field_vars   = {}
        self._modified    = False
        self._overlay_path = _asset("overlay.pdf")

        self._draw_logo()
        self._setup_ui()

        # Load first file
        if self._queue:
            self._goto(0)
        elif initial_path and os.path.exists(initial_path):
            self._load_file(initial_path)

    # ------------------------------------------------------------------
    def _draw_logo(self):
        c = tk.Canvas(self.root, height=74, bg=self.CLR_BG,
                      highlightthickness=0, bd=0)
        c.pack(fill="x")
        c.create_line(0, 73, 2000, 73, fill=self.CLR_BORDER, width=1)
        c.create_text(400, 26, text="TRAVEL  WIZARDS",
                      font=("Georgia", 26, "bold"),
                      fill="#000000", anchor="center")
        c.create_text(400, 52, text="I N V O I C E   E D I T O R",
                      font=("Arial", 10), fill="#cc0000", anchor="center")

    # ------------------------------------------------------------------
    def _setup_ui(self):
        # ── Top bar: file path + browse (single mode) / counter (queue) ─
        top = tk.Frame(self.root, bg=self.CLR_BG)
        top.pack(fill="x", padx=24, pady=(14, 0))

        if self._queue:
            # Queue mode: counter label + file name
            tk.Label(top, textvariable=self.counter_var,
                     font=("Arial", 10, "bold"),
                     bg=self.CLR_BG, fg=self.CLR_MUTED).pack(anchor="w")
            tk.Label(top, textvariable=self.xlsx_path,
                     font=("Consolas", 10),
                     bg=self.CLR_BG, fg="#888888").pack(anchor="w")
        else:
            # Single / standalone mode: browse row
            tk.Label(top, text="INVOICE FILE  (.xlsx)",
                     font=("Arial", 10, "bold"),
                     bg=self.CLR_BG, fg=self.CLR_MUTED).pack(anchor="w")
            file_box = tk.Frame(top, bg=self.CLR_PANEL,
                                highlightbackground=self.CLR_BORDER,
                                highlightthickness=1)
            file_box.pack(fill="x", pady=(4, 0))
            tk.Entry(file_box, textvariable=self.xlsx_path,
                     relief="flat", bg=self.CLR_PANEL, fg=self.CLR_TEXT,
                     readonlybackground=self.CLR_PANEL,
                     font=("Consolas", 10), bd=0,
                     state="readonly").pack(side="left", padx=10, pady=7,
                                            fill="x", expand=True)
            tk.Button(file_box, text="Browse", command=self._browse,
                      relief="flat", cursor="hand2",
                      bg=self.CLR_BTN, fg="#000000",
                      activebackground=self.CLR_BTN_ACT,
                      font=("Arial", 10, "bold"),
                      padx=14, pady=5, bd=0).pack(side="right", padx=6, pady=4)

        # ── Scrollable form ──────────────────────────────────────────────
        form_container = tk.Frame(self.root, bg=self.CLR_BG)
        form_container.pack(fill="both", expand=True, padx=24, pady=(14, 0))

        canvas = tk.Canvas(form_container, bg=self.CLR_BG,
                           highlightthickness=0, bd=0)
        sb = ttk.Scrollbar(form_container, orient="vertical",
                           command=canvas.yview)
        self.form_frame = tk.Frame(canvas, bg=self.CLR_BG)
        self.form_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.form_frame, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        # Bind scroll to this window only — unbind on close to avoid
        # "invalid command name" errors after the window is destroyed
        def _on_scroll(e):
            try:
                canvas.yview_scroll(-1 * (e.delta // 120), "units")
            except Exception:
                pass

        self.root.bind("<MouseWheel>", _on_scroll)
        self.root.bind("<Destroy>",
            lambda e: self.root.unbind("<MouseWheel>") if e.widget is self.root else None)

        self._build_form()

        # ── Status bar ───────────────────────────────────────────────────
        self.status_lbl = tk.Label(self.root, textvariable=self.status_var,
                                   font=("Arial", 9),
                                   bg=self.CLR_BG, fg=self.CLR_MUTED)
        self.status_lbl.pack(anchor="w", padx=24, pady=(8, 0))

        # ── Action buttons ───────────────────────────────────────────────
        btn_area = tk.Frame(self.root, bg=self.CLR_BG)
        btn_area.pack(fill="x", padx=24, pady=(6, 16))

        # Save .xlsx
        self.save_btn = tk.Button(
            btn_area, text="💾  Save",
            command=self._save_xlsx,
            relief="flat", cursor="hand2",
            bg=self.CLR_BTN, fg="#000000",
            activebackground=self.CLR_BTN_ACT,
            font=("Arial", 11, "bold"),
            pady=9, padx=16, bd=0)
        self.save_btn.pack(side="left", padx=(0, 8))

        # Export PDF
        self.pdf_btn = tk.Button(
            btn_area, text="📄  Export PDF",
            command=self._export_pdf,
            relief="flat", cursor="hand2",
            bg=self.CLR_BTN, fg="#000000",
            activebackground=self.CLR_BTN_ACT,
            font=("Arial", 11, "bold"),
            pady=9, padx=16, bd=0)
        self.pdf_btn.pack(side="left", padx=(0, 8))


        # Queue navigation (right-aligned)
        if self._queue:
            nav = tk.Frame(btn_area, bg=self.CLR_BG)
            nav.pack(side="right")

            self.prev_btn = tk.Button(
                nav, text="◀  Prev",
                command=self._prev,
                relief="flat", cursor="hand2",
                bg=self.CLR_BTN, fg="#000000",
                activebackground=self.CLR_BTN_ACT,
                font=("Arial", 11), pady=9, padx=14, bd=0)
            self.prev_btn.pack(side="left", padx=(0, 6))

            self.next_btn = tk.Button(
                nav, text="Next  ▶",
                command=self._next,
                relief="flat", cursor="hand2",
                bg=self.CLR_BTN, fg="#000000",
                activebackground=self.CLR_BTN_ACT,
                font=("Arial", 11), pady=9, padx=14, bd=0)
            self.next_btn.pack(side="left")

            self._update_nav_buttons()

    # ------------------------------------------------------------------
    def _build_form(self):
        for w in self.form_frame.winfo_children():
            w.destroy()
        self.field_vars.clear()

        sections = [
            ("INVOICE DETAILS", ["Invoice #", "Date", "Account", "PNR Locator"]),
            ("HOTEL",           ["Hotel Name", "Hotel Address",
                                 "Hotel City/State", "Hotel Phone"]),
            ("GUEST",           ["Confirmation #", "Guest(s)", "Arrive", "Depart"]),
            ("FINANCIALS",      ["Rate", "Total", "Commission", "Pd. To Date"]),
        ]
        field_lookup = {label: (cell, editable, is_money, multiline)
                        for label, cell, editable, is_money, multiline in FIELDS}

        for section_title, labels in sections:
            hdr = tk.Frame(self.form_frame, bg=self.CLR_SECTION,
                           highlightbackground=self.CLR_BORDER,
                           highlightthickness=1)
            hdr.pack(fill="x", pady=(12, 0))
            tk.Label(hdr, text=section_title,
                     font=("Arial", 9, "bold"),
                     bg=self.CLR_SECTION, fg=self.CLR_MUTED,
                     padx=8, pady=4).pack(anchor="w")

            for label in labels:
                cell, editable, is_money, multiline = field_lookup[label]
                var = tk.StringVar()
                self.field_vars[cell] = var

                row_frame = tk.Frame(self.form_frame, bg=self.CLR_BG)
                row_frame.pack(fill="x", pady=(4, 0))

                tk.Label(row_frame, text=label,
                         font=("Arial", 10),
                         bg=self.CLR_BG, fg=self.CLR_MUTED,
                         width=18, anchor="e").pack(side="left", padx=(0, 10))

                box = tk.Frame(row_frame,
                               highlightbackground=self.CLR_BORDER,
                               highlightthickness=1, bg=self.CLR_PANEL)
                box.pack(side="left", fill="x", expand=True, padx=(0, 4))

                if multiline:
                    txt = tk.Text(box, height=4,
                                  relief="flat", bd=0,
                                  bg=self.CLR_PANEL, fg=self.CLR_TEXT,
                                  insertbackground=self.CLR_TEXT,
                                  font=("Consolas", 11), wrap="word")
                    txt.pack(fill="both", padx=6, pady=5)
                    def _on_change(event, v=var, t=txt):
                        v.set(t.get("1.0", "end-1c"))
                        self._modified = True
                    txt.bind("<KeyRelease>", _on_change)
                    self.field_vars[cell + "_widget"] = txt
                else:
                    entry = tk.Entry(box, textvariable=var,
                                     relief="flat", bd=0,
                                     bg=self.CLR_PANEL, fg=self.CLR_TEXT,
                                     insertbackground=self.CLR_TEXT,
                                     font=("Consolas", 11))
                    entry.pack(fill="x", padx=8, pady=6)
                    var.trace_add("write", lambda *_: self._mark_modified())

    # ------------------------------------------------------------------
    def _mark_modified(self):
        self._modified = True

    def _update_nav_buttons(self):
        if not hasattr(self, "prev_btn"):
            return
        n = len(self._queue)
        self.prev_btn.config(state="normal" if self._q_index > 0     else "disabled")
        self.next_btn.config(state="normal" if self._q_index < n - 1 else "disabled")

    def _goto(self, index):
        """Navigate to queue[index], prompting to save if modified."""
        if self._modified:
            ans = messagebox.askyesnocancel(
                "Unsaved Changes",
                f"{os.path.basename(self.xlsx_path.get())} has unsaved changes.\n\n"
                "Save before continuing?")
            if ans is None:       # Cancel — stay here
                return
            if ans:               # Yes — save then navigate
                self._save_xlsx(silent=True)

        self._q_index = index
        self.counter_var.set(
            f"File {index + 1} of {len(self._queue)}")
        self._load_file(self._queue[index])
        self._update_nav_buttons()

    def _prev(self):
        if self._q_index > 0:
            self._goto(self._q_index - 1)

    def _next(self):
        if self._q_index < len(self._queue) - 1:
            self._goto(self._q_index + 1)

    # ------------------------------------------------------------------
    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select invoice .xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if path:
            self._load_file(path)

    # ------------------------------------------------------------------
    def _load_file(self, path):
        try:
            values = read_fields(path)
            self.xlsx_path.set(path)

            for label, cell, editable, is_money, multiline in FIELDS:
                var = self.field_vars.get(cell)
                txt = self.field_vars.get(cell + "_widget")
                val = values.get(cell, "")
                if multiline and txt:
                    txt.delete("1.0", "end")
                    txt.insert("1.0", val)
                    if var:
                        var.set(val)
                elif var:
                    var.set(val)

            self._modified = False
            self.status_lbl.config(fg="#1a7a1a")
            self.status_var.set(f"Loaded: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not load file:\n{e}")

    def _collect_values(self) -> dict:
        values = {}
        for label, cell, editable, is_money, multiline in FIELDS:
            txt = self.field_vars.get(cell + "_widget")
            var = self.field_vars.get(cell)
            values[cell] = txt.get("1.0", "end-1c") if (multiline and txt) \
                           else (var.get() if var else "")
        return values

    # ------------------------------------------------------------------
    def _save_xlsx(self, silent=False):
        path = self.xlsx_path.get()
        if not path:
            if not silent:
                messagebox.showerror("Error", "No file loaded.")
            return
        try:
            write_fields(path, self._collect_values(), path)
            self._modified = False
            self.status_lbl.config(fg="#1a7a1a")
            self.status_var.set(f"✓  Saved: {os.path.basename(path)}")
        except Exception as e:
            if not silent:
                messagebox.showerror("Error", f"Save failed:\n{e}")

    # ------------------------------------------------------------------
    def _resolve_overlay(self) -> str | None:
        """Return overlay path, or show a clear error if missing."""
        if os.path.exists(self._overlay_path):
            return self._overlay_path
        messagebox.showerror(
            "Overlay Not Found",
            f"overlay.pdf not found at:\n{self._overlay_path}\n\n"
            "Place overlay.pdf in the same folder as this script.")
        return None

    def _export_pdf(self):
        path = self.xlsx_path.get()
        if not path:
            messagebox.showerror("Error", "No file loaded.")
            return

        overlay = self._resolve_overlay()
        if not overlay:
            return

        # Auto-save pending changes first
        if self._modified:
            try:
                write_fields(path, self._collect_values(), path)
                self._modified = False
            except Exception as e:
                messagebox.showerror("Error", f"Could not save before export:\n{e}")
                return

        # Output folder: final_invoices/ next to processed_invoices/
        base_dir   = os.path.dirname(path)
        parent_dir = os.path.dirname(base_dir) \
                     if os.path.basename(base_dir) == "processed_invoices" \
                     else base_dir
        final_dir  = os.path.join(parent_dir, "final_invoices")
        os.makedirs(final_dir, exist_ok=True)

        # Filename: "{invoice#} {guest name}.pdf"
        vals       = self._collect_values()
        guest_raw  = vals.get("A24", "").strip()
        first_line = guest_raw.splitlines()[0].strip() if guest_raw else ""
        safe_guest = re.sub(r'[\\/*?:"<>|]', "_", first_line) if first_line \
                     else "invoice"
        invoice_no = vals.get("D3", "").strip()
        stem       = f"{invoice_no} {safe_guest}" if invoice_no else safe_guest
        out_pdf    = os.path.join(final_dir, stem + ".pdf")

        self.pdf_btn.config(state="disabled", text="⏳  Exporting…", bg="#555555")
        self.status_lbl.config(fg=self.CLR_MUTED)
        self.status_var.set("Converting to PDF…")
        self.root.update_idletasks()

        vals = self._collect_values()

        def _do():
            try:
                build_pdf(vals, overlay, out_pdf)
                self.root.after(0, lambda: self._export_done(out_pdf))
            except Exception as e:
                self.root.after(0, lambda err=e: self._export_error(err))

        threading.Thread(target=_do, daemon=True).start()

    def _export_done(self, out_pdf):
        self.pdf_btn.config(state="normal", text="📄  Export PDF",
                            bg=self.CLR_BTN, fg="#000000")
        self.status_lbl.config(fg="#1a7a1a")
        self.status_var.set(f"✓  PDF saved: {os.path.basename(out_pdf)}")
        if messagebox.askyesno("Export Complete",
                               f"PDF saved:\n{out_pdf}\n\nOpen it now?"):
            try:
                if sys.platform == "win32":   os.startfile(out_pdf)
                elif sys.platform == "darwin": subprocess.run(["open", out_pdf])
                else:                          subprocess.run(["xdg-open", out_pdf])
            except Exception:
                pass

    def _export_error(self, err):
        self.pdf_btn.config(state="normal", text="📄  Export PDF",
                            bg=self.CLR_BTN, fg="#000000")
        self.status_lbl.config(fg="#cc0000")
        self.status_var.set("Export failed.")
        messagebox.showerror("Export Failed", str(err))

    # ------------------------------------------------------------------
    def run(self):
        if self.standalone:
            self.root.mainloop()


# ---------------------------------------------------------------------------
# Standalone entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    initial = sys.argv[1] if len(sys.argv) > 1 else None
    InvoiceEditorWindow(initial_path=initial).run()
